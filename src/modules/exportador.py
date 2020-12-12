import json
from openpyxl import Workbook, utils
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border

from modules.definiciones.Materia import Materia
from modules.definiciones.Programa import Programa
from modules.definiciones.Alumno import Alumno
from modules.definiciones.Grupo import Grupo
from modules.funcionesAyuda import FuncionesAyuda as ayuda

class Exportador:
    def __init__(self, archivo):
        self.archivo = archivo

    def guardar(self, materias, alumnos, programas):
        export = {
            'alumnos': [alumno.aDict() for alumno in alumnos],
            'materias': {},
            'programas': {}
        }

        for materia in materias:
            export['materias'][materia] = materias[materia].aDict()

        for programa in programas:
            export['programas'][programa] = programas[programa].aDict()

        jsonExport = json.dumps(export, indent=2)
        with open(self.archivo, 'w') as f:
            f.write(jsonExport)

    def cargar(self):
        materias = {}
        alumnos = []
        programas = {}
        with open(self.archivo, 'r') as f:
            load = json.load(f)

        ## Cargar materias
        for materia in load['materias']:
            m = load['materias'][materia]
            materiaActual = Materia(
                m['nombre'], m['programas'], m['cuatri'],
                m['horasPorSemana'], m['tieneLab'], m['prerequisito'],
                )
            for grupo in m['grupos']:
                g = Grupo(grupo['alumnos'], materiaActual, grupo['id'])
                g.horario = grupo['horario']
                materiaActual.grupos.append(g)
            materias[materia] = materiaActual

        ## Cargar alumnos
        for alumno in load['alumnos']:
            alumnoActual = Alumno(
                alumno['registro'], alumno['materiasPendientes'],
                alumno['carrera'], alumno['materiasPorCuatri'],
                alumno['cuatri']
            )
            # poner referencias a las materias, no solo los ids
            materiasP = []
            for materia in alumnoActual.materiasPendientes:
                materiasP.append(materias[materia])
            alumnoActual.materiasPendientes = materiasP
            alumnos.append(alumnoActual)

        ## Poner referencias a obj alumno en grupos, no solo registro
        for materia in materias.values():
            for grupo in materia.grupos:
                alumnosRef = []
                for regAlumno in grupo.alumnos:
                    al = Alumno.buscarAlumno(regAlumno, alumnos)
                    alumnosRef.append(al)
                grupo.alumnos = alumnosRef

        ## Cargar programas
        for programa in load['programas'].values():
            programas[programa['nombre']] = Programa(programa['nombre'], programa['materiasPorCuatri'])

        return materias, alumnos, programas

    def exportarExcel(self, materias):
        wb = Workbook()
        ws = wb.active
        ws.title = "Grupos"

        #Encabezado
        ws.cell(row = 1, column = 1, value = 'Carreras')
        ws.cell(row = 1, column = 2, value = 'Planes')
        ws.cell(row = 1, column = 3, value = 'Materia')
        ws.cell(row = 1, column = 4, value = 'Grupo')
        ws.cell(row = 1, column = 5, value = 'Alumnos')
        ws.cell(row = 1, column = 6, value =  'L')
        ws.cell(row = 1, column = 7, value = 'M')
        ws.cell(row = 1, column = 8, value = 'Mi')
        ws.cell(row = 1, column = 9, value = 'J')
        ws.cell(row = 1, column = 10, value = 'V')

        for fila in ws.iter_rows(min_row = 1, max_col = 10, max_row = 1):
            for celda in fila:
                celda.fill = PatternFill(patternType = 'solid', start_color = '0a7cff', end_color = '0a7cff')
                celda.font = Font(name = 'Calibri', size = 13, color='FFFFFF', bold = True)
                celda.alignment = Alignment(horizontal = 'centerContinuous')

        # Grupos
        fila = 2
        for materia in materias.values():
            carreras, planes = ayuda.extraerCarreraPlan(materia.programas)
            for grupo in materia.grupos:
                ws.cell(row = fila, column = 1, value = carreras)
                ws.cell(row = fila, column = 2, value = planes)
                ws.cell(row = fila, column = 3, value = materia.nombre)
                ws.cell(row = fila, column = 4, value = grupo.id)
                ws.cell(row = fila, column = 5, value = len(grupo.alumnos))
                for i, dia in enumerate(grupo.horario.values()):
                    ws.cell(row = fila, column = 6+i, value = ', '.join([str(hora) for hora in dia]))

                fila += 1

            # Ajustar el ancho de las columnas
            for columna in ws.columns:
                maxLen = max(len(str(celda.value)) for celda in columna)
                ws.column_dimensions[utils.get_column_letter(columna[0].column)].width\
                    = maxLen if maxLen > 8 else 9
                for celda in columna:
                    celda.border = Border(top=Side(style='thin'),
                                        right=Side(style='thin'),
                                        left=Side(style='thin'),
                                        bottom=Side(style='thin'),
                                        )
        wb.save(self.archivo)
