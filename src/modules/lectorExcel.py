from openpyxl import load_workbook
from modules.definiciones.Materia import Materia
from modules.definiciones.Programa import Programa
from modules.definiciones.Alumno import Alumno
from modules.funcionesAyuda import FuncionesAyuda as ayuda
import sys
import re

class Lector:
    def __init__(self, archivoPrincipal, archivoAdicionales, mapasCurriculares):
        self._wbPrincipal = load_workbook(filename=archivoPrincipal)
        self._wbAdicional = load_workbook(filename=archivoAdicionales)
        self._mapas = self.cargarMapasCurriculares(mapasCurriculares)
        self._prerequisitos = self.extraerPrerequisitos()
        self._materiasIgnoradas = self.extraerMateriasIgnoradas()

    def cerrarArchivos(self):
        self._wbPrincipal.close()
        self._wbAdicional.close()

    def extraerPrerequisitos(self):
        '''
        Extrae prerequisitos, regresa un diccionario
        con claves con nombres de programa, que a su vez son
        diccionarios con las materias dentro, siendo la llave el
        nombre de la materia normalizado y el valor el nombre normalizado
        del prerequisito
        '''
        prerequisitos = {}
        try:
            ws = self._wbAdicional['Prerequisitos']
        except:
            print('Hoja con prerequisitos de materias no encontrado')
            sys.exit(1)

        for celda in tuple(ws.columns)[0]:
            if celda.row != 1:
                programa = ws.cell(column = 1, row = celda.row).value
                materia = ws.cell(column = 2, row = celda.row).value
                prerequisito = ws.cell(column = 3, row = celda.row).value

                if not programa in prerequisitos:
                    prerequisitos[programa] = {}

                prerequisitos[programa][ayuda.normalizar(materia)] = ayuda.normalizar(prerequisito)

        return prerequisitos

    def extraerCantidadMaterias(self, programa):
        cantMaterias = []
        try:
            ws = self._wbAdicional['MateriasPorCuatri']
        except:
            print('Hoja con cantidad de materias por cuatri no encontrado')
            sys.exit(1)

        for celda in tuple(ws.columns)[0]:
            if celda.value == programa:
                cuatri = int(ws.cell(column = celda.column+1, row = celda.row).value)
                materias = int(ws.cell(column = celda.column+2, row = celda.row).value)
                cantMaterias.insert(cuatri, materias)
        return cantMaterias

    def cargarMapasCurriculares(self, doc):
        mapas = {}
        try:
            wb = load_workbook(filename = doc)
        except:
            print('No se pudo abrir el archivo de mapas mapas curriculares')
            sys.exit(1)

        for hoja in wb:
            programa = hoja.title
            materias = {}

            for columna in hoja.columns:
                cuatri = 0
                for celda in columna:
                    if celda.row == 1:
                        cuatri = celda.value
                        continue
                    elif celda.value is None:
                        break
                    info = ayuda.normalizar(celda.value).split('#')
                    materia = {
                        'cuatri':{
                            programa: cuatri
                        },
                        'horasPorSemana': int(info[1]),
                        'tieneLab': True if len(info) == 3 else False
                    }
                    materias[ayuda.normalizar(info[0])] = materia
            mapas[programa] = materias
            wb.close()
        return mapas

    def extraerMaterias(self):
        '''
        Extrae materias de todas las hojas en el documento
        Regresa un diccionario, siendo las llaves los nombres de la materia normalizados
            (sin acentos y en minusculas)
        y los valores objetos tipo Materia
        '''
        materias = {}
        for hoja in self._wbPrincipal:
            ws = self._wbPrincipal[hoja.title]
            materiasPrograma = self._extraerMateriasWS(hoja.title, self._prerequisitos)

            # Checar si hay materias duplicadas
            # Si si hay, nos quedamos con la original y le agregamos el programa nuevo
            # y el cuatri de ese programa
            # a la lista de programas
            for materiaP in dict(materiasPrograma).keys():
                if materiaP in materias:
                    programa = materiasPrograma[materiaP].programas[0]
                    info = self.buscarMateriaEnMapas(materiasPrograma[materiaP].nombre, programa)

                    materias[materiaP].programas.extend(materiasPrograma[materiaP].programas)
                    materias[materiaP].agregarCuatri(programa, info['cuatri'])
                    materiasPrograma.pop(materiaP)
            materias.update(materiasPrograma)
        return materias

    def extraerProgramas(self):
        programas = {}
        for hoja in self._wbPrincipal:
            programa = hoja.title
            programas[programa] = Programa(programa, self.extraerCantidadMaterias(programa))
        return programas

    def extraerAlumnos(self, materias):
        '''Extrae alumnos de todas las hojas de calculo.
        Recibe un diccionario de objetos Materia (previamente extra??dos)
        Un alumno tiene no. registro, carrera y materias pendientes,
        materiasPendientes es una lista de referencias a materias del diccionario
        '''
        alumnos = []
        for hoja in self._wbPrincipal:
            for i, fila in enumerate(hoja.rows):
                if i in [0, 1]:
                     continue
                try:
                     registro = int(fila[0].value)
                     nombre = fila[1].value
                     materiasPendientes = []
                except:
                     break

                for celda in fila:
                    if celda.column in [1, 2]:
                        continue

                    try: # si ya no hay mas materias
                        nombreMateria = ayuda.extraerNombreMateria(hoja.cell(column = celda.column, row = 2).value)
                        materia = materias[ayuda.normalizar(nombreMateria)]
                    except:
                        # No se encontr?? la materia en el diccionario de materias, puede ser que sea
                        # una materia ignorada
                        continue


                    if celda.value is None:
                        calificacion = 0
                    elif type(celda.value) == int or type(celda.value) == float:
                        calificacion = celda.value
                    elif celda.value.isnumeric():
                        calificacion = int(celda.value)
                    elif celda.value == 'NI':
                        calificacion = 10
                    else:
                        # Si es un float en formato string
                        try:
                            calificacion = float(celda.value)
                        except:
                            calificacion = 0

                    if calificacion <= 5:
                        materiasPendientes.append(materia)

                alumnos.append(Alumno(registro, nombre, materiasPendientes, hoja.title))
        return alumnos

    def extraerMateriasIgnoradas(self):
        '''
        Extraer las materias que se deben ignorar de la hoja
        'MateriasIgnoradas' del archivo de datos adicionales

        Regresa una lista de regexp correspondiendo al nombre
        de cada materia normalizado (min??sculas y sin acentos ni espacios)
        '''
        materias = []
        try:
            ws = self._wbAdicional['MateriasIgnoradas']
        except:
            print('Hoja con Materias Ignoradas no encontrado')
            return []

        for celda in tuple(ws.columns)[0]:
            if celda.row != 1:
                materia = celda.value
                materia = materia.replace('*', '.+')
                materia = ayuda.normalizar(materia)
                materias.append(re.compile(materia, re.DOTALL))

        return materias


    def _extraerMateriasWS(self, programa, prerequisitos):
        '''
        Extrae las materias de una hoja de Calculo
        Regresa un diccionario, las claves siendo el nombre de la materia
        los valores siendo objetos tipo Materia
        '''
        materias = {}
        try:
            ws = self._wbPrincipal[programa]
        except:
            print(f'Error cargando hoja {programa}')
            sys.exit(1)

        for celda in tuple(ws.rows)[1]:
            if celda.column in [1, 2]:
                continue

            materia = celda.value
            try: #llave del diccionario es el nombre de la materia normalizado
                nombreMateria = ayuda.extraerNombreMateria(materia)
            except:
                break

            # Si est?? dentro de las materias ignoradas, ignorar y seguir
            if self.esMateriaIgnorada(ayuda.normalizar(nombreMateria)):
                continue

            try:
                prerequisito = prerequisitos[programa][ayuda.normalizar(nombreMateria)]
            except:
                # Si la materia tiene un numeral romano, agregar el anterior
                # como prerrequisito
                # ej. Ingles V > Ingles IV
                prerequisito = ayuda.normalizar(ayuda.materiaNumeralAnterior(nombreMateria))

            info = self.buscarMateriaEnMapas(nombreMateria, programa)

            materias[ayuda.normalizar(nombreMateria)] = \
                Materia(nombreMateria, [programa], info['cuatri'], info['horasPorSemana'], info['tieneLab'], prerequisito)

        return materias

    def buscarMateriaEnMapas(self, nombreMateria, programa):
        valorDefault = {
            'cuatri': {
                programa: 1
            },
            'horasPorSemana': 5,
            'tieneLab': False
        }
        try:
            info = self._mapas[programa][ayuda.normalizar(nombreMateria)]
        except:
            print(f'[!] No se pudo encontrar {programa}-{nombreMateria} en mapas curriculares')
            info = valorDefault

        return info

    def esMateriaIgnorada(self, nombreMateria):
        for mi in self._materiasIgnoradas:
            if mi.search(ayuda.normalizar(nombreMateria)):
                return True
        return False
